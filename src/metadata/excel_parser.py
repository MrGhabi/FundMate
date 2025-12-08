#!/usr/bin/env python3
"""Excel metadata extractor for broker statements."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Callable

import pandas as pd
from loguru import logger
from openpyxl import load_workbook


@dataclass
class ExcelMetadata:
    broker_name: Optional[str]
    account_id: Optional[str]
    statement_date: Optional[str]


class ExcelMetadataExtractor:
    def __init__(self):
        self._parsers: Dict[str, Callable[[Path], Optional[ExcelMetadata]]] = {
            "MS": self._parse_ms,
            "GS": self._parse_gs,
            "TENFUND": self._parse_tenfund,
            "TC": self._parse_trade_confirmation,
            "GSPB": self._parse_gspb,
        }

    def _load_excel_df(self, path: Path, sheet_name=0, nrows: int = 40):
        """
        Robust loader for metadata extraction (handles old .xls via xlrd3 fallback).
        """
        try:
            return pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=nrows)
        except Exception:
            pass
        try:
            import xlrd3 as xlrd
            xlrd.sheet.Sheet.handle_note = lambda self, data, txos: None  # type: ignore
            book = xlrd.open_workbook(path)
            sh = book.sheet_by_index(sheet_name if isinstance(sheet_name, int) else 0)
            rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(min(sh.nrows, nrows))]
            return pd.DataFrame(rows)
        except Exception as exc:
            logger.warning(f"Metadata excel load failed for {path}: {exc}")
            return pd.DataFrame()

    def detect_metadata(self, path: Path) -> Optional[ExcelMetadata]:
        key = self._detect_broker_key(path)

        # Parser order: hinted broker first (except TC), then other non-TC parsers, TC last.
        ordered_keys = []
        if key and key != "TC":
            ordered_keys.append(key)
        ordered_keys.extend(k for k in self._parsers if k not in ordered_keys and k != "TC")
        ordered_keys.append("TC")

        for parser_key in ordered_keys:
            parser = self._parsers[parser_key]
            try:
                result = parser(path)
            except Exception as exc:
                logger.warning(f"Excel metadata parser failed for {path} using {parser_key}: {exc}")
                continue
            if result:
                return result
        return None

    def _detect_broker_key(self, path: Path) -> Optional[str]:
        name = path.name.lower()
        parts = [p.lower() for p in path.parts]
        if any("gspb" in p or "gs pb" in p.replace(" ", "") for p in parts):
            return "GSPB"
        if "tenfund" in name or "ten fund" in name:
            return "TENFUND"
        if "gs" in name or "goldman" in name:
            return "GS"
        if "ms" in name or "morgan" in name:
            return "MS"
        if "trade confirmation" in name:
            return "TC"
        return None

    def _parse_ms(self, path: Path) -> Optional[ExcelMetadata]:
        df = self._load_excel_df(path, sheet_name=0, nrows=40)
        account = None
        stmt_date = None
        brand_hit = any(k in path.name.lower() for k in ["ms", "morgan", "optiondaily"])

        search_region = df.iloc[:20, :8] if not df.empty else pd.DataFrame()
        for value in search_region.values.flatten():
            if isinstance(value, str):
                upper = value.upper()
                if any(tag in upper for tag in ["OPTIONDAILY", "OPTION DAILY", "TEN ASSET MANAGEMENT LIMITED", "MORGAN STANLEY"]):
                    brand_hit = True
                if ("A/C" in upper or "ACCOUNT" in upper) and re.search(r"[A-Z0-9]{4,}", upper):
                    account = account or re.search(r"([A-Z0-9]{4,})", upper).group(1)
                if "VALUATION DATE" in upper or re.search(r"\d{2}-[A-Za-z]{3}-\d{4}", upper):
                    stmt_date = stmt_date or self._normalize_date(upper.split(":")[-1].strip())
            elif isinstance(value, datetime):
                stmt_date = stmt_date or value.strftime("%Y-%m-%d")

        if not (brand_hit and account and stmt_date):
            return None
        return ExcelMetadata("MORGAN STANLEY", account, stmt_date)

    def _parse_gs(self, path: Path) -> Optional[ExcelMetadata]:
        df = self._load_excel_df(path, sheet_name=0, nrows=40)
        account = None
        stmt_date = None
        brand_hit = "gs" in path.name.lower() or "goldman" in path.name.lower()

        for value in df.values.flatten():
            if isinstance(value, str):
                upper = value.upper()
                if "GOLDMAN" in upper or "GS PB" in upper:
                    brand_hit = True
                if "ACCOUNT" in upper or "A/C" in upper:
                    match = re.search(r'\(([A-Z0-9]+)\)', value)
                    if not match:
                        match = re.search(r'([A-Z0-9]{6,})', value)
                    if match:
                        account = match.group(1)
                if "VALUATION DATE" in upper or re.search(r'\d{2}-[A-Za-z]{3}-\d{4}', value):
                    stmt_date = self._normalize_date(value.split(":")[-1].strip())
            elif isinstance(value, datetime):
                stmt_date = value.strftime("%Y-%m-%d")
            elif isinstance(value, (int, float)) and not account:
                digits = str(value).replace(".0", "")
                if digits.isdigit():
                    account = digits

        if not (brand_hit and account and stmt_date):
            return None
        return ExcelMetadata("GOLDMAN SACHS", account, stmt_date)

    def _parse_tenfund(self, path: Path) -> Optional[ExcelMetadata]:
        brand_hit = "tenfund" in path.name.lower() or "ten fund" in path.name.lower()
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        stmt_date = None
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=5):
            for cell in row:
                value = cell.value
                if isinstance(value, datetime):
                    stmt_date = value.strftime("%Y-%m-%d")
                    break
                if isinstance(value, str) and "TEN FUND" in value.upper():
                    brand_hit = True
            if stmt_date:
                break
        wb.close()
        if not brand_hit:
            return None
        return ExcelMetadata("TEN FUND", None, stmt_date)

    def _parse_trade_confirmation(self, path: Path) -> Optional[ExcelMetadata]:
        # Table-shape guard: TC headers must be present to avoid misrouting.
        tc_headers = {"TRADE DATE", "BUY/SELL", "STOCK CODE", "QUANTITY", "AVG. PRICE", "AMOUNT (USD)"}
        shape_matches_tc = False
        try:
            df_head = pd.read_excel(path, sheet_name=0, nrows=5)
            cols = {str(c).strip().upper() for c in df_head.columns}
            row0 = {str(v).strip().upper() for v in df_head.iloc[0].tolist() if pd.notna(v)} if df_head.shape[0] > 0 else set()
            if len(tc_headers & cols) >= 3 or len(tc_headers & row0) >= 3:
                shape_matches_tc = True
        except Exception:
            shape_matches_tc = False

        # Minimal TC support: infer date from filename first; fall back to sheet inspection.
        name = path.name
        stmt_date = None

        # 1) Try filename patterns: 20250807 / 2025-08-07 / 2025_08_07.
        m = re.search(r"(20\d{2})[-_/]?(\d{2})[-_/]?(\d{2})", name)
        if m:
            stmt_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

        # 2) Inspect top rows for datetime or date-like strings if still unknown.
        if not stmt_date:
            df = pd.read_excel(path, sheet_name=0, header=None, nrows=30)
            for value in df.values.flatten():
                if isinstance(value, datetime):
                    stmt_date = value.strftime("%Y-%m-%d")
                    break
                if isinstance(value, str):
                    norm = self._normalize_date(value)
                    if norm and re.match(r"\d{4}-\d{2}-\d{2}", norm):
                        stmt_date = norm[:10]
                        break

        if not shape_matches_tc:
            return None
        return ExcelMetadata("TC", None, stmt_date)

    def _parse_gspb(self, path: Path) -> Optional[ExcelMetadata]:
        # Fixed-position read + fallback scan for GSPB
        broker = "GSPB"
        stmt_date = None
        account = None

        def scan_df(df):
            nonlocal stmt_date, account
            # Header scan
            for idx in range(min(df.shape[0], 12)):
                row = df.iloc[idx].tolist()
                row_upper = [str(v).upper() for v in row]
                if any("BUSINESS DATE" in cell for cell in row_upper):
                    # Try next columns in the same row for the actual date value
                    for val in row[1:]:
                        if isinstance(val, datetime):
                            stmt_date = stmt_date or val.strftime("%Y-%m-%d")
                        elif isinstance(val, str):
                            norm = self._normalize_date(val)
                            if norm and len(norm) >= 10:
                                stmt_date = stmt_date or norm[:10]
                if any("ACCOUNT NUMBER" in cell for cell in row_upper):
                    try:
                        col_idx = next(i for i, cell in enumerate(row_upper) if "ACCOUNT NUMBER" in cell)
                        acct_val = df.iloc[idx + 1, col_idx] if idx + 1 < df.shape[0] else None
                        if isinstance(acct_val, (int, float)):
                            account = account or str(int(acct_val))
                        elif isinstance(acct_val, str):
                            m = re.search(r'(\\d{6,})', acct_val.replace(",", ""))
                            account = account or (m.group(1) if m else acct_val.strip())
                    except Exception:
                        pass
                if stmt_date and account:
                    return
            # Fallback scan
            for value in df.values.flatten():
                if isinstance(value, datetime):
                    stmt_date = stmt_date or value.strftime("%Y-%m-%d")
                if isinstance(value, str):
                    upper = value.upper()
                    if "BUSINESS DATE" in upper or "DATE" in upper:
                        norm = self._normalize_date(value.split(":")[-1].strip())
                        if norm and len(norm) >= 10:
                            stmt_date = stmt_date or norm[:10]
                    if "ACCOUNT" in upper and not account:
                        m = re.search(r'(\\d{6,})', upper.replace(",", ""))
                        if m:
                            account = m.group(1)
                if stmt_date and account:
                    break

        df = self._load_excel_df(path, sheet_name=0, nrows=80)
        if df is not None and not df.empty:
            scan_df(df)
        if not stmt_date or not account:
            df2 = self._load_excel_df(path, sheet_name=1, nrows=80)
            if df2 is not None and not df2.empty:
                scan_df(df2)
        # If still missing account/date, try sibling cash/position file in same dir (pairing)
        path_obj = Path(path)
        if (not account or not stmt_date) and "Position" in path_obj.name:
            sibling = next(path_obj.parent.glob("*Cash*.xls*"), None)
            if sibling and sibling != path_obj:
                sibling_md = self._parse_gspb(sibling)
                if sibling_md:
                    account = account or sibling_md.account_id
                    stmt_date = stmt_date or sibling_md.statement_date
        elif (not account or not stmt_date) and "Cash" in path_obj.name:
            sibling = next(path_obj.parent.glob("*Position*.xls*"), None)
            if sibling and sibling != path_obj:
                sibling_md = self._parse_gspb(sibling)
                if sibling_md:
                    account = account or sibling_md.account_id
                    stmt_date = stmt_date or sibling_md.statement_date
        # Require at least brand + (account or date) to avoid misclassifying TC/others.
        name_lower = path.name.lower().replace(" ", "")
        brand_hit = "gspb" in name_lower or "gspb" in path.parent.name.lower().replace(" ", "")
        if not brand_hit or (not account and not stmt_date):
            return None
        return ExcelMetadata(broker, account, stmt_date)

    def _normalize_date(self, value: str) -> Optional[str]:
        if not value:
            return None
        value = value.strip()
        if "-" in value and len(value) >= 10:
            try:
                dt = datetime.strptime(value.strip()[:11], "%d-%b-%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return value[:10]
        if "/" in value:
            parts = value.split("/")
            if len(parts) == 3:
                if len(parts[0]) == 4:
                    return f"{parts[0]}-{parts[1]}-{parts[2][:2]}"
                return f"{parts[2][:4]}-{parts[0]}-{parts[1]}"
        if " " in value and ',' in value:
            try:
                dt = datetime.strptime(value[:12], "%b %d, %Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                pass
        return value
